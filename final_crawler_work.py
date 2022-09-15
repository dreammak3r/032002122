import requests
import re
import openpyxl
import os
import pyecharts.options as opts
from lxml import etree
from pyecharts.charts import Map3D
from pyecharts.globals import ChartType
from pyecharts.commons.utils import JsCode
from pyecharts.charts import Bar, Grid, Line, Pie, Tab
from pyecharts.charts import Line


app_issue = ['本土病例', '北京', '天津', '上海', '重庆', '河北', '山西',
             '辽宁', '吉林', '黑龙江', '江苏', '浙江', '安徽', '福建', '江西', '山东', '河南', '湖北', '湖南',
             '广东', '海南', '四川', '贵州', '云南', '陕西', '甘肃', '青海', '新疆', '内蒙古', '广西', '西藏',
             '宁夏','台湾新增','香港新增','澳门新增','台湾地区', '香港特别行政区', '澳门特别行政区']
not_issue = ['本土', '北京', '天津', '上海', '重庆', '河北', '山西', '辽宁',
             '吉林', '黑龙江', '江苏', '浙江', '安徽', '福建', '江西', '山东', '河南', '湖北', '湖南', '广东',
             '海南', '四川', '贵州', '云南', '陕西', '甘肃', '青海', '新疆', '内蒙古', '广西', '西藏', '宁夏']
headers = {
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
        'Accept-Encoding': 'gzip, deflate',
        'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
        'Cache-Control': 'max-age=0',
        'Connection': 'keep-alive',
        'Cookie': 'yfx_c_g_u_id_10006654=_ck22090517092711557123132504572; yfx_f_l_v_t_10006654=f_t_1662368967158__r_t_1662514638127__v_t_1662514638127__r_c_2; yfx_mr_10006654=%3A%3Amarket_type_free_search%3A%3A%3A%3Abaidu%3A%3A%3A%3A%3A%3A%3A%3Awww.baidu.com%3A%3A%3A%3Apmf_from_free_search; yfx_mr_f_10006654=%3A%3Amarket_type_free_search%3A%3A%3A%3Abaidu%3A%3A%3A%3A%3A%3A%3A%3Awww.baidu.com%3A%3A%3A%3Apmf_from_free_search; yfx_key_10006654=; sVoELocvxVW0S=57yh5eHi6BlWwbYuOEUHFMNXf_2SF8UL5VWS1759zdOiiImwtyLuvBL1rWffIpGlLnMEMoxnpQBHoAej5Qug.gG; security_session_verify=fc465e4f1828940ec0438b63374ada0a; sVoELocvxVW0T=53SI0.DWUeQ7qqqDkmRH3_AToYARjKiHRH568jKOM4B.OPNB2axXw5kqAtweBhHBYQOYh3hRO8OaMl8SZuRBb4HDDy8wWx_H9KnDMJfOHJhLKqwvylr_gmnhMbVf7Xl1INInmRUZl8aTrrguv1MWZmyUOXCgg2aOx6_4J72Gm.uCLEdwxtjF7hWLGGpO..CyBUuKFNGN8o.f7i5cTf3DueMgKy959yMbnxH14vnDsH.wVdK4nQbz4PLAMqYCxYgwjT4eY2xujVIScYsPVxnC5uNL45UyizBvMCagu5cjSPGfoWPa5mHqCrzryZOZ96c0axMDYQTxAdJ7LeVECT_l6vDQTiBIU5G26AzdTJ07AoQza; insert_cookie=91349450',
        'Host': 'www.nhc.gov.cn',
        'Referer': 'https://cn.bing.com/',
        'Upgrade-Insecure-Requests': '1',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/105.0.0.0 Safari/537.36 Edg/105.0.1343.33'}

#创建条形图
def bar_create(bar_time_data, bar_covid_data, bar_not_app_covid_data):
    bar_time_data.reverse()
    bar_covid_data.reverse()
    bar_not_app_covid_data.reverse()
    c = (
        Bar(init_opts=opts.InitOpts(width="1600px", height="900px", theme="essos"))
        .add_xaxis(bar_time_data)
        .add_yaxis("新增普通确诊", bar_covid_data)
        .add_yaxis("新增无症状确诊", bar_not_app_covid_data)
        .set_global_opts(
            title_opts=opts.TitleOpts(title="最近24天新增确诊病例"),
            datazoom_opts=[opts.DataZoomOpts()],
        )
    )
    return c

#创建折线图
def line_create(line_time_data, line_covid_data, line_not_app_covid_data):

    line_time_data.reverse()
    line_covid_data.reverse()
    line_not_app_covid_data.reverse()
    c = (
        Line(init_opts=opts.InitOpts(width="1600px", height="900px", theme="essos"))
        .set_global_opts(
            tooltip_opts=opts.TooltipOpts(is_show=True),
            xaxis_opts=opts.AxisOpts(type_="category"),
            yaxis_opts=opts.AxisOpts(
                type_="value",
                axistick_opts=opts.AxisTickOpts(is_show=True),
                splitline_opts=opts.SplitLineOpts(is_show=True),
            ),
            datazoom_opts=[opts.DataZoomOpts()]
        )
        .add_xaxis(xaxis_data=line_time_data)
        .add_yaxis(
            series_name="新增普通确诊",
            y_axis=line_covid_data,
            symbol="emptyCircle",
            is_symbol_show=True,
            label_opts=opts.LabelOpts(is_show=True),linestyle_opts=opts.LineStyleOpts(width=2)
        )
        .add_yaxis(
            series_name="新增无症状确诊",
            y_axis=line_not_app_covid_data,
            symbol="emptyCircle",
            is_symbol_show=True,
            label_opts=opts.LabelOpts(is_show=True),linestyle_opts=opts.LineStyleOpts(width=2)
        )
    )
    return c

#创建3D地图
def map_3d_creat(adata, ndata):
        today_covid_data = [
            ("黑龙江", [127.9688, 45.368, adata["黑龙江"]]),
            ("内蒙古", [110.3467, 41.4899, adata["内蒙古"]]),
            ("吉林", [125.8154, 44.2584, adata["吉林"]]),
            ("辽宁", [123.1238, 42.1216, adata["辽宁"]]),
            ("河北", [114.4995, 38.1006, adata["河北"]]),
            ("天津", [117.4219, 39.4189, adata["天津"]]),
            ("山西", [112.3352, 37.9413, adata["山西"]]),
            ("陕西", [109.1162, 34.2004, adata["陕西"]]),
            ("甘肃", [103.5901, 36.3043, adata["甘肃"]]),
            ("宁夏", [106.3586, 38.1775, adata["宁夏"]]),
            ("青海", [101.4038, 36.8207, adata["青海"]]),
            ("新疆", [87.9236, 43.5883, adata["新疆"]]),
            ("西藏", [91.11, 29.97, adata["西藏"]]),
            ("四川", [103.9526, 30.7617, adata["四川"]]),
            ("重庆", [108.384366, 30.439702, adata["重庆"]]),
            ("山东", [117.1582, 36.8701, adata["山东"]]),
            ("河南", [113.4668, 34.6234, adata["河南"]]),
            ("江苏", [118.8062, 31.9208, adata["江苏"]]),
            ("安徽", [117.29, 32.0581, adata["安徽"]]),
            ("湖北", [114.3896, 30.6628, adata["湖北"]]),
            ("浙江", [119.5313, 29.8773, adata["浙江"]]),
            ("福建", [119.4543, 25.9222, adata["福建"]]),
            ("江西", [116.0046, 28.6633, adata["江西"]]),
            ("湖南", [113.0823, 28.2568, adata["湖南"]]),
            ("贵州", [106.6992, 26.7682, adata["贵州"]]),
            ("广西", [108.479, 23.1152, adata["广西"]]),
            ("海南", [110.3893, 19.8516, adata["海南"]]),
            ("上海", [121.4648, 31.2891, adata["上海"]]),
            ("北京", [116.46, 39.92, adata["北京"]]),
            ("香港", [114.1733, 22.3200, adata["香港新增"]]),
            ("云南", [100.2969, 25.7217, adata["云南"]]),
            ("澳门", [111.9586, 21.8, adata["澳门新增"]]),
            ("台湾", [121.5200, 25.0307, adata["台湾新增"]])
        ]
        today_not_app_covid_data = [
            ("黑龙江", [127.9688, 45.368, ndata["黑龙江"]]),
            ("内蒙古", [110.3467, 41.4899, ndata["内蒙古"]]),
            ("吉林", [125.8154, 44.2584, ndata["吉林"]]),
            ("辽宁", [123.1238, 42.1216, ndata["辽宁"]]),
            ("河北", [114.4995, 38.1006, ndata["河北"]]),
            ("天津", [117.4219, 39.4189, ndata["天津"]]),
            ("山西", [112.3352, 37.9413, ndata["山西"]]),
            ("陕西", [109.1162, 34.2004, ndata["陕西"]]),
            ("甘肃", [103.5901, 36.3043, ndata["甘肃"]]),
            ("宁夏", [106.3586, 38.1775, ndata["宁夏"]]),
            ("青海", [101.4038, 36.8207, ndata["青海"]]),
            ("新疆", [87.9236, 43.5883, ndata["新疆"]]),
            ("西藏", [91.11, 29.97, ndata["西藏"]]),
            ("四川", [103.9526, 30.7617, ndata["四川"]]),
            ("重庆", [108.384366, 30.439702, ndata["重庆"]]),
            ("山东", [117.1582, 36.8701, ndata["山东"]]),
            ("河南", [113.4668, 34.6234, ndata["河南"]]),
            ("江苏", [118.8062, 31.9208, ndata["江苏"]]),
            ("安徽", [117.29, 32.0581, ndata["安徽"]]),
            ("湖北", [114.3896, 30.6628, ndata["湖北"]]),
            ("浙江", [119.5313, 29.8773, ndata["浙江"]]),
            ("福建", [119.4543, 25.9222, ndata["福建"]]),
            ("江西", [116.0046, 28.6633, ndata["江西"]]),
            ("湖南", [113.0823, 28.2568, ndata["湖南"]]),
            ("贵州", [106.6992, 26.7682, ndata["贵州"]]),
            ("广西", [108.479, 23.1152, ndata["广西"]]),
            ("海南", [110.3893, 19.8516, ndata["海南"]]),
            ("上海", [121.4648, 31.2891, ndata["上海"]]),
            ("北京", [116.46, 39.92, ndata["北京"]]),
            ("云南", [100.2969, 25.7217, ndata["云南"]])
        ]
        map = (
            Map3D(init_opts=opts.InitOpts(width="1600px", height="900px", ))
            .add_schema(
                box_height = 50,
                region_height=1,
                view_control_opts=opts.Map3DViewControlOpts(distance=150, pan_mouse_button='middle', rotate_mouse_button='left'),
                itemstyle_opts=opts.ItemStyleOpts(
                    color="rgb(17,63,61)",
                    opacity=1,
                    border_width=0.8,
                    border_color="rgb(35,235,185)",
                ),
                map3d_label=opts.Map3DLabelOpts(
                    is_show=False,
                    formatter=JsCode("function(data){return data.name + " " + data.value[2];}"),
                ),
                emphasis_label_opts=opts.LabelOpts(
                    is_show=False,
                    color="#fff",
                    font_size=10,
                    background_color="rgba(0,23,11,0)",
                ),
                light_opts=opts.Map3DLightOpts(
                    main_color="#fff",
                    main_intensity=1.2,
                    main_shadow_quality="high",
                    is_main_shadow=True,
                    main_beta=10,
                    ambient_intensity=0.3,
                ),

            )
            .add(
                series_name="今日新增确诊",
                data_pair=today_covid_data,
                type_=ChartType.BAR3D,
                shading="lambert",
                min_height=0.1,
                label_opts=opts.LabelOpts(
                    is_show=False,
                    formatter=JsCode("function(data){return data.name + ' ' + data.value[2];}"),
                ),
            )
            .add(
                series_name="今日新增无症状确诊",
                data_pair=today_not_app_covid_data,
                type_=ChartType.BAR3D,
                bar_size=1,
                shading="lambert",
                min_height=0.1,
                label_opts=opts.LabelOpts(
                    is_show=False,
                    formatter=JsCode("function(data){return data.name + ' ' + data.value[2];}"),
                ),
            )
            .set_global_opts(title_opts=opts.TitleOpts(title="今日疫情数据"),
                             visualmap_opts=opts.VisualMapOpts(max_=20, range_color=[
                                 "#33CC00",
                                 "#3EB804",
                                 "#4AA308",
                                 "#558F0B",
                                 "#617A0F",
                                 "#6C6613",
                                 "#775217",
                                 "#833D1B",
                                 "#8E291E",
                                 "#9A1422",
                                 "#A50026",
                             ], ), tooltip_opts=opts.TooltipOpts(is_show= False), legend_opts=opts.LegendOpts(selected_mode='single'))
        )
        return map

#初始化excel表格
def init_xls():
    wb = openpyxl.Workbook()
    del wb["Sheet"]
    sheet_a = wb.create_sheet("新增普通")
    sheet_n = wb.create_sheet("新增无症状")

    sheet_a.cell(1, 1, "日期")
    for x1, x2 in zip(range(2, 40), app_issue):
        sheet_a.cell(1, x1, x2)
    # 初始化本土无症状新增表格头
    sheet_n.cell(1, 1, "日期")
    for x1, x2 in zip(range(2, 34), not_issue):
        sheet_n.cell(1, x1, x2)

    return wb

#写入excel表格中
def write_into_xls(workbook, time_data_final, i, count_line, len_html, adata, ndata, range_up_limit):
    #获取表格工作表
    sheet_a = workbook["新增普通"]
    sheet_n = workbook["新增无症状"]

    #计算该写入的行数
    row = 1 + (i - 1) * 24 + count_line + 1

    #普通新增数据写入
    sheet_a.cell(row, 1).value = time_data_final#写入普通新增的日期
    for x1 in range(2, 40):
        #台湾，香港，澳门特殊处理，写入公式用累计新增计算新增
        if sheet_a.cell(1, x1).value == "台湾新增":
            if (i == range_up_limit - 1 and count_line == len_html - 1):
                sheet_a.cell(row, x1).value = 0
            else:#字符串拼接成公式
                str1 = 'AK' + str(row)
                str2 = 'AK' + str(row+1)
                str3 = str1 + '-' + str2
                sheet_a.cell(row, x1).value = '=' + str3

        elif sheet_a.cell(1, x1).value == "香港新增":
            if (i == range_up_limit - 1 and count_line == len_html - 1):
                sheet_a.cell(row, x1).value = 0
            else:
                str1 = 'AL' + str(row)
                str2 = 'AL' + str(row+1)
                str3 = str1 + '-' + str2
                sheet_a.cell(row, x1).value = '=' + str3

        elif sheet_a.cell(1, x1).value == "澳门新增":
            if (i == range_up_limit - 1 and count_line == len_html - 1):
                sheet_a.cell(row, x1).value = 0
            else:
                str1 = 'AM' + str(row)
                str2 = 'AM' + str(row+1)
                str3 = str1 + '-' + str2
                sheet_a.cell(row, x1).value = '=' + str3

        elif (sheet_a.cell(1, x1).value) in adata:
            sheet_a.cell(row, x1).value = adata[sheet_a.cell(1, x1).value]


    #无症状新增数据写入
    sheet_n.cell(row, 1).value = time_data_final
    for x1 in range(2, 34):
        sheet_n.cell(row, x1).value = ndata[sheet_n.cell(1, x1).value]

    #保存excel文件
    workbook.save("covid-19.xls")

#请求网页并解析文本
def crwal_and_text(range_up_limit, workbook):
    # 柱状图所需数据定义
    bar_time_data = []
    bar_covid_data = []
    bar_not_app_covid_data = []

    #爬取网页
    for i in range(1, range_up_limit):
        if i == 1:
            url = 'http://www.nhc.gov.cn/xcs/yqtb/list_gzbd.shtml'
        else:
            url = 'http://www.nhc.gov.cn/xcs/yqtb/list_gzbd_{}.shtml'.format(i)

        req = requests.get(url, headers=headers)
        e = etree.HTML(req.text)
        html = e.xpath('//div[@class="w1024 mb50"]/div[@class="list"]/ul[@class="zxxx_list"]/li/a/@href')
        title = e.xpath('//div[@class="w1024 mb50"]/div[@class="list"]/ul[@class="zxxx_list"]/li/a/@title')


        #折线图所需数据定义
        line_time_data=[]
        line_covid_data=[]
        line_not_app_covid_data = []




        for html_ite, title_c, count_line in zip(html, title, range(len(html))):

            #省份转换成字典
            adata = dict.fromkeys(app_issue, 0)
            ndata = dict.fromkeys(not_issue, 0)

            # 访问网站并解析html
            surl = "http://www.nhc.gov.cn" + html_ite
            resq = requests.get(surl, headers=headers)
            se = etree.HTML(resq.text)

            # 获取纯文本信息
            words_num = se.xpath('string(.)')
            text = ''.join(words_num)

            # 单独获取新增确诊和新增无症状港澳台三段文字且进行拼接
            # 正则表达式模式
            yiqing_pattern = '[\u4e00-\u9fa5]{2,}[0-9]+(?=例|人)'
            first_pattern = re.compile('新增确诊[\u4e00-\u9fa5]+[0-9]+例[。|，][^。]+|新增无症状[\u4e00-\u9fa5]+[0-9]+例[。|，][^。]+|累计收到[\u4e00-\u9fa5]+[0-9]+[^ ]+')
            digit_pattern = '[0-9]+'
            hanzi_pattern = '[\u4e00-\u9fa5]+'
            date_pattern = '[0-9]+月[0-9]+日'
            in_in_pattern = '本土病例[0-9]+例[^。）]+|本土[0-9]+例[^。）]+|累计收到[\u4e00-\u9fa5]+[0-9]+[^ ]+'

            #获取日期信息
            time_data = re.findall(date_pattern, title_c)
            time_data_final = ''.join(time_data)

            # 文本处理
            says_1 = re.findall(first_pattern, text)
            temp_1 = ''.join(says_1)
            says_2 = re.findall(in_in_pattern, temp_1)
            temp_2 = '\n'.join(says_2)
            final_says = re.findall(yiqing_pattern, temp_2)

            # 病例信息录入字典
            flag = 1
            for item in final_says:
                words = ''.join(re.findall(hanzi_pattern, item))
                num = ''.join(re.findall(digit_pattern, item))

                #文本解析后天然顺序普通新增在前面无症状在中间港澳台在最后所有设置一个flag来区分录入的信息
                if words == "本土病例":#普通新增flag=1
                    flag = 1
                elif words == "本土":#无症状flag=0
                    flag = 0

                if flag == 1:
                    if (words in adata):
                        adata[words] = int(num)
                    else:
                        continue

                elif flag == 0:
                    if (words in ndata):
                        ndata[words] = int(num)

                    elif (words in adata):#港澳台同一录入普通新增
                        adata[words] = int(num)
                    else:
                        continue

            print(time_data_final + "信息开始录入")

            #柱状图所需要的数据
            bar_time_data.append(time_data_final)
            bar_covid_data.append(adata["本土病例"])
            bar_not_app_covid_data.append(ndata["本土"])

            #收集各类图表所需要的数据
            line_time_data.append(time_data_final)
            line_covid_data.append(adata["本土病例"])
            line_not_app_covid_data.append(ndata["本土"])

            # 3d地图所需数据
            if (count_line == 0 and i == 1):
                today_covid_adata = adata
                today_not_app_covid_adata = ndata

            if (count_line == len(html) - 1):
                creat_all_charts(today_covid_adata, today_not_app_covid_adata, line_time_data, line_covid_data, line_not_app_covid_data, bar_time_data,
                                 bar_covid_data, bar_not_app_covid_data)

            write_into_xls(workbook, time_data_final, i, count_line, len(html), adata, ndata, range_up_limit)
        #控制台检测数据#爬取

#控制台输出函数默认关闭
# def data_consle_print(data, title):
#     print(title+':')
#     for key, value in zip(data.keys(), data.values()):
#         print(key, ":", value)
#     print("======================================================")

        #将疫情信息导入表格中

#创建所有图表
def creat_all_charts(today_covid_adata, today_not_app_covid_adata, line_time_data, line_covid_data, line_not_app_covid_data, bar_time_data, bar_covid_data, bar_not_app_covid_data):
    #分页创建所有图表
    tab = Tab()
    tab.add(map_3d_creat(today_covid_adata, today_not_app_covid_adata), "今日确诊地图")
    tab.add(line_create(line_time_data, line_covid_data, line_not_app_covid_data), "24天确诊折线图")
    tab.add(bar_create(bar_time_data, bar_covid_data, bar_not_app_covid_data), "24天确诊柱状图")
    tab.render("mycharts.html")
    os.system("mycharts.html")

if __name__ == "__main__":
    workbook = init_xls()
    page=int(input("请输入您要爬取的页数（一页24天）："))
    crwal_and_text(page+1, workbook)
